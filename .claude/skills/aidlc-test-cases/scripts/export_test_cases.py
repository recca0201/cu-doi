#!/usr/bin/env python3
"""Export AIDLC test-case Markdown to CSV and/or Excel."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import asdict, dataclass
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = None


DEFAULT_STANDARD_SCHEMA = [
    {"header": "Test Case ID", "field": "test_case_id"},
    {"header": "Title", "field": "title"},
    {"header": "Source ID", "field": "source_id"},
    {"header": "Source Summary", "field": "source_summary"},
    {"header": "Coverage Type", "field": "coverage_type"},
    {"header": "Automation Target", "field": "automation_target"},
    {"header": "Priority", "field": "priority"},
    {"header": "Preconditions", "field": "preconditions"},
    {"header": "Test Steps", "field": "test_steps"},
    {"header": "Expected Result", "field": "expected_result"},
    {"header": "Tags", "field": "tags"},
    {"header": "Notes", "field": "notes"},
]

DEFAULT_XRAY_SCHEMA = [
    {"header": "External ID", "field": "test_case_id"},
    {"header": "Summary", "field": "title"},
    {"header": "Issue Type", "value": "Test"},
    {"header": "Test Type", "value": "Manual"},
    {"header": "Priority", "field": "priority"},
    {"header": "Labels", "field": "tags"},
    {"header": "Description", "field": "xray_description"},
    {"header": "Manual Test Step", "field": "test_steps"},
    {"header": "Manual Test Data", "field": "preconditions"},
    {"header": "Manual Test Result", "field": "expected_result"},
]

DEFAULT_XLSX_WIDTHS = {
    "A": 16,
    "B": 42,
    "C": 18,
    "D": 36,
    "E": 18,
    "F": 18,
    "G": 12,
    "H": 36,
    "I": 36,
    "J": 36,
    "K": 24,
    "L": 28,
}


@dataclass
class ScenarioData:
    """Parsed Gherkin scenario content."""

    title: str
    preconditions: str
    test_steps: str
    expected_result: str


@dataclass
class ExportRow:
    """Normalized row shape used by all export profiles."""

    test_case_id: str
    title: str
    source_id: str
    source_summary: str
    coverage_type: str
    automation_target: str
    priority: str
    preconditions: str
    test_steps: str
    expected_result: str
    tags: str
    notes: str


@dataclass
class TraceabilityRow:
    """Traceability mapping from source IDs to summaries."""

    source_ids: set[str]
    source_summary: str


def load_export_schema() -> dict:
    """Load export schema metadata from the skill assets."""
    schema_path = Path(__file__).resolve().parents[1] / "assets" / "export-schema.json"
    if not schema_path.exists():
        return {
            "standard": DEFAULT_STANDARD_SCHEMA,
            "xray": DEFAULT_XRAY_SCHEMA,
            "xlsx_widths": DEFAULT_XLSX_WIDTHS,
        }

    parsed = json.loads(schema_path.read_text(encoding="utf-8"))
    return {
        "standard": parsed.get("standard", DEFAULT_STANDARD_SCHEMA),
        "xray": parsed.get("xray", DEFAULT_XRAY_SCHEMA),
        "xlsx_widths": parsed.get("xlsx_widths", DEFAULT_XLSX_WIDTHS),
    }


def normalize_header(value: str) -> str:
    """Normalize a Markdown table header for matching."""
    return re.sub(r"[^a-z0-9]+", " ", value.strip().lower()).strip()


def split_table_row(line: str) -> list[str]:
    """Split a simple Markdown table row."""
    return [part.strip() for part in line.strip().strip("|").split("|")]


def is_separator_row(line: str) -> bool:
    """Return whether a Markdown table row is the header separator."""
    parts = split_table_row(line)
    return bool(parts) and all(re.fullmatch(r":?-{3,}:?", part.strip()) for part in parts)


class AidlcMarkdownParser:
    """Parse supported AIDLC test-case Markdown formats."""

    def __init__(self, content: str):
        self.content = content

    def parse(self) -> list[ExportRow]:
        """Parse rows from gherkin-table, standard-table, or gherkin-blocks."""
        rows = self._parse_gherkin_table()
        if rows:
            return rows

        rows = self._parse_standard_table()
        if rows:
            return rows

        return self._parse_gherkin_blocks()

    def _extract_section(self, heading: str, next_headings: list[str]) -> str:
        start_marker = f"## {heading}"
        start = self.content.find(start_marker)
        if start == -1:
            return ""

        after_start = start + len(start_marker)
        end_positions = []
        for candidate in next_headings:
            pos = self.content.find(f"## {candidate}", after_start)
            if pos != -1:
                end_positions.append(pos)

        end = min(end_positions) if end_positions else len(self.content)
        return self.content[start:end]

    def _parse_markdown_tables(self, section: str) -> list[list[dict[str, str]]]:
        lines = [line.strip() for line in section.splitlines()]
        tables: list[list[dict[str, str]]] = []
        index = 0

        while index < len(lines) - 1:
            if not lines[index].startswith("|") or not is_separator_row(lines[index + 1]):
                index += 1
                continue

            headers = split_table_row(lines[index])
            normalized_headers = [normalize_header(header) for header in headers]
            index += 2
            rows: list[dict[str, str]] = []

            while index < len(lines) and lines[index].startswith("|"):
                values = split_table_row(lines[index])
                row = {
                    header: values[pos] if pos < len(values) else ""
                    for pos, header in enumerate(normalized_headers)
                }
                rows.append(row)
                index += 1

            if rows:
                tables.append(rows)

        return tables

    def _parse_traceability_matrix(self) -> list[TraceabilityRow]:
        section = self._extract_section(
            "Traceability matrix",
            ["Test cases", "Scenario inventory", "Gherkin scenarios", "Coverage summary", "Coverage gaps"],
        )
        if not section:
            return []

        mappings: list[TraceabilityRow] = []
        for table in self._parse_markdown_tables(section):
            for row in table:
                source_id = row.get("source id", "")
                source_summary = row.get("source summary", "")
                if source_id:
                    mappings.append(
                        TraceabilityRow(
                            source_ids=self._split_source_ids(source_id),
                            source_summary=source_summary,
                        )
                    )
        return mappings

    def _resolve_source_summary(
        self,
        source_id_text: str,
        traceability: list[TraceabilityRow],
    ) -> str:
        wanted = self._split_source_ids(source_id_text)
        if not wanted:
            return ""

        for row in traceability:
            if wanted.issubset(row.source_ids):
                return row.source_summary

        for row in traceability:
            if wanted & row.source_ids:
                return row.source_summary

        return ""

    def _split_source_ids(self, source_id_text: str) -> set[str]:
        return {part.strip() for part in source_id_text.split(",") if part.strip()}

    def _build_tags(self, coverage_type: str, automation_target: str, priority: str) -> str:
        tags = [
            coverage_type.replace(",", " "),
            automation_target,
            priority.lower(),
        ]
        return ", ".join(tag for tag in tags if tag)

    def _normalize_export_row(
        self,
        test_case_id: str,
        title: str,
        source_id: str,
        coverage_type: str,
        automation_target: str,
        priority: str,
        preconditions: str,
        test_steps: str,
        expected_result: str,
        notes: str,
        traceability: list[TraceabilityRow],
    ) -> ExportRow | None:
        if not test_case_id.strip() or not title.strip():
            return None

        return ExportRow(
            test_case_id=test_case_id.strip(),
            title=title.strip(),
            source_id=source_id.strip(),
            source_summary=self._resolve_source_summary(source_id, traceability),
            coverage_type=coverage_type.strip(),
            automation_target=automation_target.strip(),
            priority=priority.strip(),
            preconditions=preconditions.strip(),
            test_steps=test_steps.strip(),
            expected_result=expected_result.strip(),
            tags=self._build_tags(coverage_type, automation_target, priority),
            notes=notes.strip(),
        )

    def _parse_gherkin_table(self) -> list[ExportRow]:
        section = self._extract_section(
            "Test cases",
            ["Coverage summary", "Coverage gaps", "Automation notes", "Review checklist"],
        )
        traceability = self._parse_traceability_matrix()
        rows: list[ExportRow] = []

        for table in self._parse_markdown_tables(section):
            if not table:
                continue
            required = {"test case id", "scenario", "given", "when", "then", "source id"}
            if not required.issubset(table[0].keys()):
                continue

            for item in table:
                row = self._normalize_export_row(
                    test_case_id=item.get("test case id", ""),
                    title=item.get("scenario", ""),
                    source_id=item.get("source id", ""),
                    coverage_type=item.get("coverage type", ""),
                    automation_target=item.get("automation target", ""),
                    priority=item.get("priority", ""),
                    preconditions=item.get("given", ""),
                    test_steps=item.get("when", ""),
                    expected_result=item.get("then", ""),
                    notes=item.get("notes", ""),
                    traceability=traceability,
                )
                if row:
                    rows.append(row)

        return rows

    def _parse_standard_table(self) -> list[ExportRow]:
        section = self._extract_section(
            "Test cases",
            ["Coverage summary", "Coverage gaps", "Automation notes", "Review checklist"],
        )
        traceability = self._parse_traceability_matrix()
        rows: list[ExportRow] = []

        for table in self._parse_markdown_tables(section):
            if not table:
                continue
            required = {"test case id", "title", "preconditions", "steps", "expected result"}
            if not required.issubset(table[0].keys()):
                continue

            for item in table:
                row = self._normalize_export_row(
                    test_case_id=item.get("test case id", ""),
                    title=item.get("title", ""),
                    source_id=item.get("source id", ""),
                    coverage_type=item.get("coverage type", ""),
                    automation_target=item.get("automation target", ""),
                    priority=item.get("priority", ""),
                    preconditions=item.get("preconditions", ""),
                    test_steps=item.get("steps", ""),
                    expected_result=item.get("expected result", ""),
                    notes=item.get("notes", ""),
                    traceability=traceability,
                )
                if row:
                    rows.append(row)

        return rows

    def _parse_gherkin_blocks(self) -> list[ExportRow]:
        traceability = self._parse_traceability_matrix()
        inventory = self._parse_scenario_inventory()
        scenarios = self._parse_gherkin_scenarios()
        scenarios_by_title = {scenario.title: scenario for scenario in scenarios}

        rows: list[ExportRow] = []
        missing = []

        for index, item in enumerate(inventory):
            title = item.get("title", "")
            scenario = scenarios_by_title.get(title)
            if not scenario:
                scenario = scenarios[index] if index < len(scenarios) else ScenarioData(title, "", "", "")
                if not scenario.preconditions and not scenario.test_steps and not scenario.expected_result:
                    missing.append(title)

            row = self._normalize_export_row(
                test_case_id=item.get("test case id", ""),
                title=title,
                source_id=item.get("source id", ""),
                coverage_type=item.get("coverage type", ""),
                automation_target=item.get("automation target", ""),
                priority=item.get("priority", ""),
                preconditions=scenario.preconditions,
                test_steps=scenario.test_steps,
                expected_result=scenario.expected_result,
                notes=item.get("notes", ""),
                traceability=traceability,
            )
            if row:
                rows.append(row)

        if missing:
            print(
                "Warning: some inventory rows had no matching Gherkin scenario title:\n- "
                + "\n- ".join(missing),
                file=sys.stderr,
            )

        return rows

    def _parse_scenario_inventory(self) -> list[dict[str, str]]:
        section = self._extract_section(
            "Scenario inventory",
            ["Gherkin scenarios", "Coverage gaps", "Automation notes"],
        )
        if not section:
            return []

        for table in self._parse_markdown_tables(section):
            if table and {"test case id", "title"}.issubset(table[0].keys()):
                return table

        raise ValueError("Scenario inventory table not found in Markdown source")

    def _parse_gherkin_scenarios(self) -> list[ScenarioData]:
        section = self._extract_section(
            "Gherkin scenarios",
            ["Coverage summary", "Coverage gaps", "Automation notes", "Review checklist"],
        )
        if not section:
            return []

        pattern = re.compile(r"```gherkin\s*\n(.*?)```", re.DOTALL)

        scenarios: list[ScenarioData] = []
        for match in pattern.finditer(section):
            block = match.group(1).strip()
            parsed = self._parse_single_scenario(block)
            if parsed:
                scenarios.append(parsed)

        return scenarios

    def _parse_single_scenario(self, block: str) -> ScenarioData | None:
        lines = [line.rstrip() for line in block.splitlines() if line.strip()]
        if not lines:
            return None

        first = lines[0].strip()
        if not first.startswith("Scenario:"):
            return None

        title = first.split(":", 1)[1].strip()
        given_steps: list[str] = []
        when_steps: list[str] = []
        then_steps: list[str] = []
        current = None

        for raw in lines[1:]:
            line = raw.strip()
            if line.startswith("Given "):
                current = "given"
                given_steps.append(line[6:].strip())
            elif line.startswith("When "):
                current = "when"
                when_steps.append(line[5:].strip())
            elif line.startswith("Then "):
                current = "then"
                then_steps.append(line[5:].strip())
            elif line.startswith("And ") or line.startswith("But "):
                text = line[4:].strip()
                if current == "given":
                    given_steps.append(text)
                elif current == "when":
                    when_steps.append(text)
                elif current == "then":
                    then_steps.append(text)

        return ScenarioData(
            title=title,
            preconditions="\n".join(given_steps),
            test_steps="\n".join(when_steps),
            expected_result="\n".join(then_steps),
        )


def build_xray_description(row: ExportRow) -> str:
    """Build the Jira/Xray CSV description field."""
    parts = [
        f"Source ID: {row.source_id}",
        f"Coverage Type: {row.coverage_type}",
        f"Automation Target: {row.automation_target}",
    ]
    if row.source_summary:
        parts.insert(1, f"Source Summary: {row.source_summary}")
    if row.notes:
        parts.append(f"Notes: {row.notes}")
    return "\n".join(parts)


def resolve_schema_value(row: ExportRow, schema_item: dict) -> str:
    """Resolve a schema item into an exported cell value."""
    if "value" in schema_item:
        return str(schema_item["value"])

    field = schema_item.get("field")
    if field == "xray_description":
        return build_xray_description(row)

    values = asdict(row)
    value = values.get(field, "")
    return "" if value is None else str(value)


def export_csv_with_profile(rows: list[ExportRow], output_path: Path, profile: str) -> None:
    """Export normalized rows to CSV."""
    schema = load_export_schema()
    columns = schema["xray"] if profile == "xray" else schema["standard"]

    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow([item["header"] for item in columns])
        for row in rows:
            writer.writerow([resolve_schema_value(row, item) for item in columns])


def export_csv(rows: list[ExportRow], output_path: Path) -> None:
    """Export normalized rows to standard CSV."""
    export_csv_with_profile(rows, output_path, profile="standard")


def export_xlsx(rows: list[ExportRow], output_path: Path) -> None:
    """Export normalized rows to Excel."""
    if Workbook is None:
        raise RuntimeError(
            "openpyxl is required for Excel export. Install it with: pip3 install openpyxl"
        )

    schema = load_export_schema()
    columns = schema["standard"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for index, item in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=index, value=item["header"])
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_index, row in enumerate(rows, start=2):
        for col_index, item in enumerate(columns, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=resolve_schema_value(row, item))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if row.priority == "Critical":
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            elif row.priority == "High":
                cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")

    for column, width in schema["xlsx_widths"].items():
        ws.column_dimensions[column].width = width

    ws.row_dimensions[1].height = 28
    for row_index in range(2, ws.max_row + 1):
        ws.row_dimensions[row_index].height = 54

    ws.freeze_panes = "A2"
    wb.save(output_path)


def default_output_paths(input_path: Path) -> tuple[Path, Path]:
    """Return default CSV and XLSX paths for a Markdown source."""
    return input_path.with_suffix(".csv"), input_path.with_suffix(".xlsx")


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(description="Export AIDLC test-case Markdown to CSV and/or Excel")
    parser.add_argument("input_markdown", help="Path to the Markdown test-case document")
    parser.add_argument("--csv", dest="csv_path", help="Path to output CSV file")
    parser.add_argument("--xlsx", dest="xlsx_path", help="Path to output Excel file")
    parser.add_argument(
        "--profile",
        choices=["standard", "xray"],
        default="standard",
        help="CSV export profile. 'xray' emits a Jira/Xray-friendly CSV shape.",
    )
    return parser.parse_args()


def main() -> int:
    """Run the export command."""
    args = parse_args()
    input_path = Path(args.input_markdown)
    if not input_path.exists():
        print(f"Error: input file not found: {input_path}", file=sys.stderr)
        return 1

    csv_path, xlsx_path = default_output_paths(input_path)
    if args.profile == "xray" and not args.csv_path:
        csv_path = input_path.with_suffix(".xray.csv")
    if args.csv_path:
        csv_path = Path(args.csv_path)
    if args.xlsx_path:
        xlsx_path = Path(args.xlsx_path)

    if not args.csv_path and not args.xlsx_path:
        args.csv_path = str(csv_path)
        args.xlsx_path = str(xlsx_path)

    content = input_path.read_text(encoding="utf-8")
    parser = AidlcMarkdownParser(content)
    rows = parser.parse()
    if not rows:
        print("Error: no exportable test-case rows found in Markdown", file=sys.stderr)
        return 1

    if args.csv_path:
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        export_csv_with_profile(rows, csv_path, profile=args.profile)
        print(f"Wrote CSV: {csv_path}")

    if args.xlsx_path:
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        export_xlsx(rows, xlsx_path)
        print(f"Wrote Excel: {xlsx_path}")

    print(f"Exported {len(rows)} test cases from {input_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
